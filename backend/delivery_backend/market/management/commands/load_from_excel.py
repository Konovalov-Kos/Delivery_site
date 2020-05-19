from django.core.management.base import BaseCommand, CommandError
import os

from market.models import Category, Product

from delivery_backend.settings import DATA_DIR

from openpyxl import load_workbook

class Command(BaseCommand):

    def handle(self, *args, **options):
        print('Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print(f'Start importing from excel {DATA_DIR}')
        wb = load_workbook(os.path.join(DATA_DIR, 'price.xlsx'))
        print(wb.get_sheet_names()[0])
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        single_category = None

        for row in range(1, sheet.max_row+1):
            item = sheet.cell(row=row, column=3).value
            id = sheet.cell(row=row, column=2).value
            if id == None:
                print('Create a new category')
                single_category = Category()
                single_category.name = item
                single_category.save()
            else:
                print('Create a new good')
                if single_category:
                    single_product = Product()
                    single_product.name = item
                    single_product.category = single_category
                    single_product.save()